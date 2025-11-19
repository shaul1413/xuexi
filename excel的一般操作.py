# #用python 实现 读入、写入、修改数据的方法
#
# #读取excel数据，xlrd
# import xlrd
# from openpyxl.cell import WriteOnlyCell
#
# book = xlrd.open_workbook(r'D:\xuexi1\book.xls')
# sheet = book.sheet_by_index(0)
# print(sheet.cell(row=1, column=2).value)#读取单元格第1行，第二列的值
#
# #写入excel数据，xlwt,只能针对xls的文件
# import xlwt
# book = xlwt.Workbook()
# sheet = book.add_sheet('sheet1')
# sheet.write(0,0,'数据')
# book.save(r'D:/xuexi1/book1.xls')
#
# #写入excel数据，openpyxl,针对xlsx文件
# import openpyxl
# wb = openpyxl.Workbook()
# sheet = wb.active
# sheet1 = wb.create_sheet('sheet1')
# sheet.cell(row=1,column=1,value='内容')
# wb.save('D:/xuexi1/book2.xlsx')
#
#
# #用openpyxl修改数据
# import openpyxl as xl
# wb = xl.Workbook()
# ws = wb.active
# wb.save(r'D:/xuexi1/book2.xlsx')
# wb = xl.load_workbook(r'D:\xuexi1\book2.xlsx')
# print(wb.sheetnames)
# ws1 = wb.active
# ws1.sheet_properties.tabColor = 'FF0000'
# ws1.cell(row=1,column=1,value='内容')
# wb.save(r'D:/xuexi1/book2.xlsx')
#
#
# #openpyxl修改表格的样式，要用到openpyxl的styles函数
# import openpyxl as xl
# from openpyxl.styles import Font,colors,Alignment
# wb = xl.load_workbook(r'D:\xuexi1\book2.xlsx')
# ws2 = wb.create_sheet('sheet1')
#
# #写入数据
# rows = [
#     ['id','name','age'],
#     ['0001','shulu',20]
# ]
# for row in rows:
#     ws2.append(row)
#
# #设置单元格字体和格式
# font = Font(name='Arial', size=12, italic=False, bold=True, color=colors.BLUE)
# ws2['A1'].font = font
# alignment = Alignment(horizontal='center', vertical='center')
# ws2['A1'].alignment = alignment
#
# #针对第五行修改行列高
# ws2.row_dimensions[5].height = 40
# ws2.column_dimensions['A'].width = 40
#
# #合并单元格
# ws2.merge_cells('A7:B7')
# ws2['A7'] = '合并后的单元格'
#
# #保存单元格
# wb.save(r'D:\xuexi1\book2.xlsx')


# #openpyxl 读取百万量级数据
# from openpyxl import load_workbook
# wb = load_workbook(r'D:\xuexi1\book2.xlsx',read_only=True)
# ws3 = wb['sheet1']
#
# for row in ws3.rows:
#     for cell in row:
#         print(cell.value)

#利用openpyxl 写入百万数量级数据到excel
import openpyxl as xl
from openpyxl.cell import WriteOnlyCell
from openpyxl.comments import Comment
from openpyxl.styles import Font

wb = xl.Workbook(write_only=True)
ws = wb.create_sheet('sheet1')
#写入表格用 WriteOnlyCell
cell = WriteOnlyCell(ws,value='shulu')
#设置单元格的格式
cell.font = Font(name='Arial', size=12)
#插入批注
cell.comment=Comment(text='这是一个批注',author='人名')
ws.append([cell,3.14,None])
wb.save(r'D:\xuexi1\book2.xlsx')